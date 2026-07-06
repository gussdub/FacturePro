"""Tests — import de relevé XLSX dans le rapprochement (feature #7.2).

Le XLSX passe par le MÊME pipeline de mapping que le CSV (_map_bank_rows) : mêmes
règles de signe, parse_error, dédup. Cible la conversion des cellules typées (dates,
nombres) et le routage endpoint.
"""
import io
import json
import os
import sys
from datetime import datetime as dt

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))
os.environ.setdefault("MONGO_URL", "mongodb://localhost:27017")
os.environ.setdefault("JWT_SECRET", "test")
os.environ.setdefault("DB_NAME", "facturepro")

import pytest  # noqa: E402
import openpyxl  # noqa: E402
from fastapi.testclient import TestClient  # noqa: E402
from backend.server import app, db, _parse_xlsx_rows, _xlsx_cell_to_str  # noqa: E402

client = TestClient(app)
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MAP_DC = {  # débit + crédit
    "delimiter": ",", "has_header": True, "date_column": 0, "date_format": "YYYY-MM-DD",
    "description_column": 1, "amount_mode": "debit_credit", "debit_column": 2, "credit_column": 3,
    "sign_convention": "positive_is_credit",
}
MAP_SINGLE = {
    "delimiter": ",", "has_header": True, "date_column": 0, "date_format": "YYYY-MM-DD",
    "description_column": 1, "amount_mode": "single", "amount_column": 2,
    "sign_convention": "positive_is_credit",
}


def _make_xlsx(rows, header=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    if header:
        ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────── UNITAIRE ───────────────────────────

def test_cell_to_str_types():
    assert _xlsx_cell_to_str(None) == ""
    assert _xlsx_cell_to_str(dt(2026, 6, 1)) == "2026-06-01"
    assert _xlsx_cell_to_str(9.20) == "9.2"       # ré-parsé par _normalize_amount
    assert _xlsx_cell_to_str(107.57) == "107.57"
    assert _xlsx_cell_to_str(5) == "5"
    assert _xlsx_cell_to_str("IONOS") == "IONOS"


def test_xlsx_debit_credit_signs():
    xlsx = _make_xlsx(header=["Date", "Desc", "Debit", "Credit"], rows=[
        [dt(2026, 6, 1), "IONOS", 9.20, None],     # achat -> débit -> négatif
        [dt(2026, 6, 2), "DEPOT", None, 100.0],    # dépôt -> crédit -> positif
    ])
    rows = _parse_xlsx_rows(xlsx, MAP_DC)
    assert rows[0]["date"] == "2026-06-01" and rows[0]["amount_cad"] == -9.20 and rows[0]["parse_error"] is False
    assert rows[1]["date"] == "2026-06-02" and rows[1]["amount_cad"] == 100.0 and rows[1]["parse_error"] is False


def test_xlsx_single_mode_and_text_date():
    # Date stockée en TEXTE dans le XLSX (pas cellule date) -> parsée selon le format.
    xlsx = _make_xlsx(header=["Date", "Desc", "Montant"], rows=[
        ["2026/06/05", "RENDER", -107.57],         # slashes tolérées (fix précédent)
    ])
    rows = _parse_xlsx_rows(xlsx, MAP_SINGLE)
    assert rows[0]["date"] == "2026-06-05"
    assert rows[0]["amount_cad"] == -107.57
    assert rows[0]["parse_error"] is False


def test_xlsx_empty_sheet_is_safe():
    xlsx = _make_xlsx(header=None, rows=[])
    assert _parse_xlsx_rows(xlsx, MAP_DC) == []


def test_xlsx_ragged_rows_no_false_mismatch():
    # Lignes de largeurs différentes (cellules vides de fin rognées par openpyxl) ne doivent
    # PAS déclencher un col_mismatch : _parse_xlsx_rows uniformise la largeur.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Desc", "Debit", "Credit"])
    ws.append([dt(2026, 6, 1), "IONOS", 9.20])     # 3 cellules (Credit manquant)
    ws.append([dt(2026, 6, 2), "DEPOT", None, 100.0])
    buf = io.BytesIO(); wb.save(buf)
    rows = _parse_xlsx_rows(buf.getvalue(), MAP_DC)
    assert all(r["parse_error"] is False for r in rows), [r for r in rows if r["parse_error"]]


def test_corrupt_xlsx_raises_valueerror():
    with pytest.raises(ValueError):
        _parse_xlsx_rows(b"PK\x03\x04 not really a workbook", MAP_DC)


def test_date_iso_fallback_for_xlsx_real_dates():
    # Une vraie date XLSX est émise en ISO ; le repli doit la parser même si l'utilisateur
    # a choisi DD/MM/YYYY ou MM/DD/YYYY -> plus de lignes rouges à tort.
    from backend.server import _parse_csv_date
    assert _parse_csv_date("2026-06-01", "DD/MM/YYYY") == "2026-06-01"
    assert _parse_csv_date("2026-06-01", "MM/DD/YYYY") == "2026-06-01"
    # un vrai DD/MM reste interprété selon le format choisi (pas de faux repli ISO)
    assert _parse_csv_date("01/06/2026", "DD/MM/YYYY") == "2026-06-01"
    assert _parse_csv_date("06/01/2026", "MM/DD/YYYY") == "2026-06-01"
    assert _parse_csv_date("pas une date", "YYYY-MM-DD") is None


def test_xlsx_real_date_cells_with_ddmmyyyy_mapping():
    # Cellules DATE réelles + mapping DD/MM/YYYY : grâce au repli ISO, les lignes ne sont
    # plus toutes en parse_error.
    xlsx = _make_xlsx(header=["Date", "Desc", "Debit", "Credit"], rows=[
        [dt(2026, 6, 1), "IONOS", 9.20, None],
        [dt(2026, 6, 2), "GITHUB", 5.74, None],
    ])
    mapping = dict(MAP_DC, date_format="DD/MM/YYYY")
    rows = _parse_xlsx_rows(xlsx, mapping)
    assert all(r["parse_error"] is False for r in rows)
    assert rows[0]["date"] == "2026-06-01" and rows[1]["date"] == "2026-06-02"


def test_xlsx_far_cell_is_bounded_and_data_intact():
    # Une cellule très à droite (col 250) ne doit ni casser ni gonfler la sortie : les colonnes
    # mappées (0-3) restent lues, la ligne de données parse correctement.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Date", "Desc", "Debit", "Credit"])
    ws.append([dt(2026, 6, 1), "IONOS", 9.20, None])
    ws.cell(row=2, column=250, value="junk lointain")  # étend la ligne de données à 250 colonnes
    buf = io.BytesIO(); wb.save(buf)
    rows = _parse_xlsx_rows(buf.getvalue(), MAP_DC)
    ionos = [r for r in rows if r["description"] == "IONOS"]
    assert len(ionos) == 1
    assert ionos[0]["amount_cad"] == -9.20 and ionos[0]["parse_error"] is False


# ─────────────────────────── INTÉGRATION ───────────────────────────

@pytest.fixture
def auth_headers():
    r = client.post("/api/auth/login", json={"email": "gussdub@gmail.com", "password": "testpass123"})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def test_xlsx_dry_run_source_and_rows(auth_headers):
    xlsx = _make_xlsx(header=["Date", "Desc", "Debit", "Credit"], rows=[
        [dt(2026, 6, 1), "IONOS", 9.20, None],
        [dt(2026, 6, 2), "GITHUB", 5.74, None],
    ])
    r = client.post("/api/bank/imports?dry_run=true", headers=auth_headers,
                    files={"file": ("releve.xlsx", xlsx, XLSX_MIME)},
                    data={"mapping": json.dumps(MAP_DC), "bank_label": "Test XLSX"})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["source"] == "xlsx"
    assert body["total_rows"] == 2
    assert sorted(row["amount_cad"] for row in body["parsed_rows"]) == [-9.20, -5.74]


def test_xlsx_full_import_and_dedup(auth_headers):
    xlsx = _make_xlsx(header=["Date", "Desc", "Debit", "Credit"], rows=[
        [dt(2026, 7, 1), "XLSXTEST VENDOR", 42.42, None],
    ])
    from backend.server import _compute_file_hash
    fh = _compute_file_hash(xlsx)
    db.bank_imports.delete_many({"file_hash": fh})
    imported_id = None
    try:
        r = client.post("/api/bank/imports", headers=auth_headers,
                        files={"file": ("releve.xlsx", xlsx, XLSX_MIME)},
                        data={"mapping": json.dumps(MAP_DC), "bank_label": "Test XLSX"})
        assert r.status_code in (200, 201), r.text
        data = r.json()
        imported_id = data["import"]["id"]
        assert data["import"]["source"] == "xlsx"
        assert len(data["transactions"]) == 1
        assert data["transactions"][0]["amount_cad"] == -42.42
        # ré-import du même XLSX -> 409
        r2 = client.post("/api/bank/imports", headers=auth_headers,
                         files={"file": ("releve.xlsx", xlsx, XLSX_MIME)},
                         data={"mapping": json.dumps(MAP_DC), "bank_label": "Test XLSX"})
        assert r2.status_code == 409, r2.text
    finally:
        if imported_id:
            db.bank_transactions.delete_many({"import_id": imported_id})
            db.bank_imports.delete_one({"id": imported_id})
